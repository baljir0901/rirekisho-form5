from flask import Flask, request, jsonify, send_file, render_template
from datetime import datetime
import openpyxl
from openpyxl.styles import Alignment
import os
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import smtplib
import traceback
from pathlib import Path
from config import EMAIL_CONFIG, FILE_CONFIG, REQUIRED_FIELDS

# Move the validate_template function definition before app initialization
def validate_template(template_path):
    """Validate Excel template structure"""
    try:
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Required cell ranges to check
        required_ranges = [
            'E5:L5', 'M5:T5',  # Name fields
            'E7:K7', 'O7:T7',  # Birth and age
            'E8:K8',           # Gender
            'E9:K9', 'O9:T9'  # Nationality and station
        ]
        
        # Check if all ranges exist
        for cell_range in required_ranges:
            try:
                _ = ws[cell_range]
            except Exception as e:
                raise ValueError(f"Template validation failed: Missing range {cell_range}")
        
        return True
        
    except Exception as e:
        print(f"Template validation error: {e}")
        raise

app = Flask(__name__)

# Validate template on startup
try:
    validate_template(FILE_CONFIG["TEMPLATE_PATH"])
    print("Template validation successful")
except Exception as e:
    print(f"Template validation failed: {e}")
    raise

def write_to_cell(ws, cell_range, value):
    """Write value to cell(s) with proper alignment"""
    try:
        # Handle None or empty values
        if value is None or value == '':
            value = ""
            
        if ':' in cell_range:
            # For merged cells
            first_cell = cell_range.split(':')[0]
            ws[first_cell] = value
            for row in ws[cell_range]:
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # For single cells, need to check if it's part of a merged range
            cell = ws[cell_range]
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                # Find the master cell of the merged range
                for merged_range in ws.merged_cells.ranges:
                    if cell_range in merged_range:
                        master_cell = ws[merged_range.start_cell]
                        master_cell.value = value
                        master_cell.alignment = Alignment(horizontal='center', vertical='center')
                        break
            else:
                # Normal cell
                cell.value = value
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
        print(f"Successfully wrote '{value}' to cell {cell_range}")  # Debug print
            
    except Exception as e:
        print(f"Error writing to cell {cell_range}: {e}")
        print(f"Value type: {type(value)}")
        print(f"Value: {value}")
        raise

def format_japanese_date(date_str):
    """Convert date string to Japanese format"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        year = date_obj.year
        month = date_obj.month
        day = date_obj.day
        return f"{year}年{month}月{day}日"
    except:
        return date_str

def fill_excel_template(form_data):
    try:
        template_path = "rirekisho_template.xlsx"
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        
        # Basic Information
        basic_info_mappings = {
            'E5:L5': form_data.get('furigana_surname'),
            'M5:T5': form_data.get('furigana_given'),
            'E7:K7': format_japanese_date(form_data.get('birthdate')),
            'O7:T7': form_data.get('age'),
            'E8:K8': form_data.get('gender'),
            'E9:K9': form_data.get('nationality'),
            'O9:T9': form_data.get('nearest_station'),
            'F10:T10': form_data.get('postal_code'),
            'E11:T11': form_data.get('address'),
            'E13:H13': form_data.get('phone'),
            'I13:N13': form_data.get('email')
        }
        
        # Interview Dates
        interview_mappings = {
            'F14:I14': form_data.get('interview_date_1'),
            'K14:N14': form_data.get('interview_date_2'),
            'P14:T14': form_data.get('interview_date_3')
        }
        
        # Write Basic Information
        for cell_range, value in basic_info_mappings.items():
            try:
                write_to_cell(ws, cell_range, value)
            except Exception as e:
                print(f"Error writing basic info to {cell_range}: {e}")
                continue

        # Write Interview Dates
        for cell_range, value in interview_mappings.items():
            try:
                write_to_cell(ws, cell_range, value)
            except Exception as e:
                print(f"Error writing interview date to {cell_range}: {e}")
                continue

        # Education History
        education_base_row = 18
        for edu in form_data.get('education', []):
            write_to_cell(ws, f'A{education_base_row}:D{education_base_row}', edu.get('entrance_date'))
            write_to_cell(ws, f'D{education_base_row}:F{education_base_row}', edu.get('graduation_date'))
            write_to_cell(ws, f'G{education_base_row}:L{education_base_row}', edu.get('school_name'))
            write_to_cell(ws, f'M{education_base_row}:T{education_base_row}', edu.get('department'))
            education_base_row += 1
        
        # Work History
        work_base_row = 24
        for work in form_data.get('work_history', []):
            write_to_cell(ws, f'A{work_base_row}:C{work_base_row}', work.get('start_date'))
            write_to_cell(ws, f'D{work_base_row}:F{work_base_row}', work.get('end_date'))
            write_to_cell(ws, f'G{work_base_row}:L{work_base_row}', work.get('company_name'))
            write_to_cell(ws, f'M{work_base_row}:T{work_base_row}', work.get('job_description'))
            work_base_row += 1

        # Visa Information
        write_to_cell(ws, 'A36:J36', form_data.get('visa_status'))
        write_to_cell(ws, 'K36:T36', form_data.get('visa_expiry'))

        # Technical Intern History
        intern_base_row = 39
        for history in form_data.get('intern_history', []):
            write_to_cell(ws, f'A{intern_base_row}', history.get('year'))
            write_to_cell(ws, f'C{intern_base_row}', history.get('month'))
            write_to_cell(ws, f'D{intern_base_row}:G{intern_base_row}', history.get('visa_status'))
            write_to_cell(ws, f'H{intern_base_row}:M{intern_base_row}', history.get('institution'))
            write_to_cell(ws, f'N{intern_base_row}:T{intern_base_row}', history.get('management_org'))
            intern_base_row += 1

        # Family in Japan
        has_family = form_data.get('has_family_in_japan', False)
        write_to_cell(ws, 'H43', '☑' if has_family else '□')
        write_to_cell(ws, 'J43', '□' if has_family else '☑')

        if has_family:
            lives_together = form_data.get('lives_together', False)
            write_to_cell(ws, 'H44', '☑' if lives_together else '□')
            write_to_cell(ws, 'J44', '□' if lives_together else '☑')

            family_base_row = 46
            for member in form_data.get('family_info', []):
                write_to_cell(ws, f'A{family_base_row}:B{family_base_row}', member.get('relation'))
                write_to_cell(ws, f'C{family_base_row}:I{family_base_row}', member.get('name'))
                write_to_cell(ws, f'J{family_base_row}:M{family_base_row}', member.get('birthdate'))
                write_to_cell(ws, f'N{family_base_row}:T{family_base_row}', member.get('workplace'))
                family_base_row += 1

        # Home Country Information
        home_base_row = 51
        for contact in form_data.get('home_country_info', []):
            write_to_cell(ws, f'A{home_base_row}:B{home_base_row}', contact.get('relation'))
            write_to_cell(ws, f'C{home_base_row}:I{home_base_row}', contact.get('name'))
            write_to_cell(ws, f'J{home_base_row}:M{home_base_row}', contact.get('address'))
            write_to_cell(ws, f'N{home_base_row}:T{home_base_row}', contact.get('contact'))
            home_base_row += 1

        # Allergies
        has_allergies = form_data.get('has_allergies', False)
        write_to_cell(ws, 'A59:F59', '☑' if has_allergies else '□')
        write_to_cell(ws, 'F59:K59', '□' if has_allergies else '☑')

        # Write allergy explanation
        write_to_cell(ws, 'A57:T57', '※有のある場合、何のアレルギーを持っているのかをチェックし、名前を書いて下さい。')

        if has_allergies:
            # Direct cell writing for allergies
            allergy_mappings = {
                'allergy_medicine': ('A58', form_data.get('allergy_medicine', False)),
                'allergy_food': ('B58', form_data.get('allergy_food', False)),
                'allergy_metal': ('C58', form_data.get('allergy_metal', False)),
                'allergy_plant': ('D58', form_data.get('allergy_plant', False)),
                'allergy_other': ('E58', form_data.get('allergy_other', False))
            }
            
            for allergy_name, (cell_ref, value) in allergy_mappings.items():
                try:
                    # Write directly to the cell reference
                    ws[cell_ref] = '☑' if value else '□'
                    ws[cell_ref].alignment = Alignment(horizontal='center', vertical='center')
                    print(f"Successfully wrote allergy '{allergy_name}' to {cell_ref}")
                except Exception as e:
                    print(f"Error writing {allergy_name} to {cell_ref}: {e}")
                    continue

            # Write allergy name label and value
            write_to_cell(ws, 'A62:F62', 'アレルギーの名前:')
            if form_data.get('allergy_names'):
                write_to_cell(ws, 'G62:Q62', form_data.get('allergy_names'))

        # Self PR
        write_to_cell(ws, 'A68:T71', form_data.get('self_pr'))

        # Physical Information
        physical_info_mappings = {
            'A74:D74': form_data.get('height'),
            'E74:H74': form_data.get('weight'),
            'I74:K74': form_data.get('shoe_size'),
            'M74': form_data.get('clothes_size_top'),
            'P74': form_data.get('clothes_size_bottom'),
            'R74:T74': form_data.get('blood_type'),
            'O75:Q75': form_data.get('waist')
        }

        for cells, value in physical_info_mappings.items():
            write_to_cell(ws, cells, value)

        # Save file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"rirekisho_{timestamp}.xlsx"
        output_path = os.path.join(FILE_CONFIG["OUTPUT_DIR"], output_filename)
        wb.save(output_path)
        return output_filename

    except Exception as e:
        print(f"Error in fill_excel_template: {e}")
        raise

def send_email(form_data, excel_file):
    try:
        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG["SENDER_EMAIL"]
        msg['To'] = EMAIL_CONFIG["RECIPIENT_EMAIL"]
        msg['Subject'] = f"新しい履歴書が提出されました - {form_data.get('furigana_surname')} {form_data.get('furigana_given')}"

        # Create email body
        body = f"""
        新しい履歴書が提出されました。

        基本情報:
        - 氏名: {form_data.get('furigana_surname')} {form_data.get('furigana_given')}
        - ロー字: {form_data.get('romaji_surname')} {form_data.get('romaji_given')}
        - 生年月日: {form_data.get('birthdate')}
        - 年齢: {form_data.get('age')}
        - 性別: {form_data.get('gender')}
        - 国籍: {form_data.get('nationality')}
        - 最寄り駅: {form_data.get('nearest_station')}

        連絡先:
        - 電話番号: {form_data.get('phone')}
        - メール: {form_data.get('email')}
        - 住所: {form_data.get('postal_code')} {form_data.get('address')}

        面接希望日:
        1. {form_data.get('interview_date_1')}
        2. {form_data.get('interview_date_2')}
        3. {form_data.get('interview_date_3')}

        在留資格: {form_data.get('visa_status')}
        在留期限: {form_data.get('visa_expiry')}

        提出日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}

        詳細は添付の Excel ファイルをご確認ください。
        """

        msg.attach(MIMEText(body, 'plain', 'utf-8'))

        # Attach Excel file
        try:
            with open(excel_file, 'rb') as f:
                excel_attachment = MIMEApplication(f.read(), _subtype='xlsx')
                excel_attachment.add_header(
                    'Content-Disposition', 
                    'attachment', 
                    filename=os.path.basename(excel_file)
                )
                msg.attach(excel_attachment)
        except Exception as e:
            print(f"Error attaching file: {e}")
            raise

        # Send email
        with smtplib.SMTP(EMAIL_CONFIG["SMTP_SERVER"], EMAIL_CONFIG["SMTP_PORT"]) as server:
            server.starttls()
            server.login(EMAIL_CONFIG["SENDER_EMAIL"], EMAIL_CONFIG["SENDER_PASSWORD"])
            server.send_message(msg)

        print(f"Email sent successfully to {EMAIL_CONFIG['RECIPIENT_EMAIL']}")
        return True

    except Exception as e:
        print(f"Error sending email: {e}")
        raise

def cleanup_old_files(directory=".", pattern="rirekisho_*.xlsx", keep_last=5):
    """Cleanup old Excel files, keeping only the most recent ones"""
    try:
        files = sorted(Path(directory).glob(pattern), key=lambda x: x.stat().st_mtime, reverse=True)
        for file in files[keep_last:]:
            try:
                file.unlink()
                print(f"Deleted old file: {file}")
            except Exception as e:
                print(f"Error deleting {file}: {e}")
    except Exception as e:
        print(f"Error in cleanup: {e}")

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit-form', methods=['POST'])
def submit_form():
    output_file = None
    try:
        print("Received request")
        form_data = request.get_json()
        print("Form data:", form_data)
        
        # Generate Excel file
        print("Generating Excel file...")
        output_file = fill_excel_template(form_data)
        print(f"Excel file generated: {output_file}")
        
        # Send email with attachment
        print("Sending email...")
        send_email(form_data, output_file)
        print("Email sent successfully")
        
        # Cleanup old files
        cleanup_old_files()
        
        return jsonify({
            "status": "success",
            "message": "履歴書が正常に送信されました"
        })
        
    except Exception as e:
        print(f"Error in submit_form: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        
        # Cleanup failed file if it exists
        if output_file and Path(output_file).exists():
            try:
                Path(output_file).unlink()
                print(f"Cleaned up failed file: {output_file}")
            except Exception as cleanup_error:
                print(f"Error cleaning up file: {cleanup_error}")
        
        return jsonify({
            "status": "error",
            "message": "送信中にエラーが発生しました。もう一度お試しください。"
        }), 500

@app.route('/test-excel', methods=['GET'])
def test_excel():
    try:
        # More comprehensive test data
        test_data = {
            'furigana_surname': 'テスト',
            'furigana_given': '太郎',
            'romaji_surname': 'TEST',
            'romaji_given': 'TARO',
            'birthdate': '2000-01-01',
            'age': '24',
            'gender': '男',
            'nationality': '日本',
            'nearest_station': '東京駅',
            'postal_code': '100-0001',
            'address': '東京都千代田区',
            'phone': '090-1234-5678',
            'email': 'test@example.com',
            'interview_date_1': '2024-01-01',
            'interview_date_2': '2024-01-02',
            'interview_date_3': '2024-01-03',
            'visa_status': '技能実習',
            'visa_expiry': '2025-01-01',
            'has_allergies': True,
            'allergy_medicine': True,
            'allergy_food': False,
            'allergy_metal': True,
            'allergy_plant': False,
            'allergy_other': False,
            'self_pr': 'テストデータです。よろしくお願いいたします。',
            'height': '170',
            'weight': '65',
            'shoe_size': '26',
            'clothes_size_top': 'M',
            'clothes_size_bottom': 'L',
            'blood_type': 'A',
            'waist': '80'
        }
        
        print("Starting test Excel generation...")
        output_file = fill_excel_template(test_data)
        
        return jsonify({
            "status": "success",
            "message": f"Test Excel file generated: {output_file}",
            "data": test_data
        })
        
    except Exception as e:
        print(f"Error in test-excel: {e}")
        print(f"Traceback: {traceback.format_exc()}")
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

@app.route('/download/<filename>')
def download_file(filename):
    try:
        file_path = os.path.join(FILE_CONFIG["OUTPUT_DIR"], filename)
        if os.path.exists(file_path):
            return send_file(
                file_path,
                as_attachment=True,
                download_name=filename
            )
        else:
            return jsonify({
                "status": "error",
                "message": "ファイルが見つかりません"
            }), 404
    except Exception as e:
        print(f"Error downloading file: {e}")
        return jsonify({
            "status": "error",
            "message": "ダウンロード中にエラーが発生しました"
        }), 500

if __name__ == '__main__':
    app.run(debug=True)